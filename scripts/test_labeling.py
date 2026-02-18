"""라벨링 테스트 스크립트 - 기존 로직 그대로 수행하되, 노션/리포트 발송만 제외하고 엑셀 파일을 Slack으로 전송"""
import sys
import os
from datetime import datetime
from typing import List, Dict, Any
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.bigquery.client import BigQueryClient
from src.bigquery.queries import WeeklyDataQuery
from src.classifier.vector_classifier import VectorContentClassifier, ServiceCategoryReviewer
from src.vectorstore.chroma_store import ChromaVectorStore
from src.reporter.analytics import WeeklyAnalytics
from src.reporter.report_generator import ReportGenerator
from src.storage.data_store import ClassifiedDataStore
from src.integrations.slack_client import SlackNotifier


# 서비스 공지글 필터링 키워드
FILTER_KEYWORDS = [
    "channel.io",
    "어떤 채팅방도 운영하지 않습니다",
    "어떤 채팅방 · 밴드도 운영하지 않습니다",
    "문의하기:",
]


def filter_service_notices(items: List[Dict[str, Any]], content_field: str = "message") -> List[Dict[str, Any]]:
    """서비스 공지글 필터링"""
    filtered = []
    removed_count = 0

    for item in items:
        content = item.get(content_field, "") or ""
        is_notice = any(keyword in content for keyword in FILTER_KEYWORDS)

        if is_notice:
            removed_count += 1
        else:
            filtered.append(item)

    if removed_count > 0:
        print(f"  ⚠️  서비스 공지글 {removed_count}건 필터링됨")

    return filtered


def create_labeled_excel(letters: List[Dict], posts: List[Dict], output_path: str) -> str:
    """분류된 데이터를 엑셀 파일로 생성"""
    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # === 편지 시트 ===
    ws_letters = wb.active
    ws_letters.title = "편지"

    letter_headers = ["마스터", "오피셜클럽", "내용", "카테고리", "생성일"]
    for col, header in enumerate(letter_headers, 1):
        cell = ws_letters.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(letters, 2):
        classification = item.get('classification', {})
        for col in range(1, 6):
            cell = ws_letters.cell(row=row_idx, column=col)
            cell.alignment = wrap_alignment

        ws_letters.cell(row=row_idx, column=1, value=item.get('masterName', ''))
        ws_letters.cell(row=row_idx, column=2, value=item.get('masterClubName', ''))
        ws_letters.cell(row=row_idx, column=3, value=(item.get('message', '') or '')[:1000])
        ws_letters.cell(row=row_idx, column=4, value=classification.get('category', ''))
        ws_letters.cell(row=row_idx, column=5, value=item.get('createdAt', '')[:10] if item.get('createdAt') else '')

    ws_letters.column_dimensions['A'].width = 12
    ws_letters.column_dimensions['B'].width = 15
    ws_letters.column_dimensions['C'].width = 80
    ws_letters.column_dimensions['D'].width = 15
    ws_letters.column_dimensions['E'].width = 12

    # === 게시글 시트 ===
    ws_posts = wb.create_sheet(title="게시글")

    post_headers = ["마스터", "오피셜클럽", "제목", "내용", "카테고리", "생성일"]
    for col, header in enumerate(post_headers, 1):
        cell = ws_posts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, item in enumerate(posts, 2):
        classification = item.get('classification', {})
        content = item.get('textBody') or item.get('body', '') or ''
        for col in range(1, 7):
            cell = ws_posts.cell(row=row_idx, column=col)
            cell.alignment = wrap_alignment

        ws_posts.cell(row=row_idx, column=1, value=item.get('masterName', ''))
        ws_posts.cell(row=row_idx, column=2, value=item.get('masterClubName', ''))
        ws_posts.cell(row=row_idx, column=3, value=(item.get('title', '') or '')[:200])
        ws_posts.cell(row=row_idx, column=4, value=content[:1000])
        ws_posts.cell(row=row_idx, column=5, value=classification.get('category', ''))
        ws_posts.cell(row=row_idx, column=6, value=item.get('createdAt', '')[:10] if item.get('createdAt') else '')

    ws_posts.column_dimensions['A'].width = 12
    ws_posts.column_dimensions['B'].width = 15
    ws_posts.column_dimensions['C'].width = 40
    ws_posts.column_dimensions['D'].width = 80
    ws_posts.column_dimensions['E'].width = 15
    ws_posts.column_dimensions['F'].width = 12

    wb.save(output_path)
    print(f"✓ 엑셀 파일 생성: {output_path}")

    return output_path


def main():
    """라벨링 테스트 메인"""
    print("=" * 60)
    print("📊 라벨링 테스트 - 분류 후 엑셀 파일 Slack 전송")
    print("=" * 60)
    print()

    # 날짜 입력
    start_date = input("시작일 (YYYY-MM-DD): ").strip()
    end_date = input("종료일 (YYYY-MM-DD): ").strip()

    try:
        datetime.strptime(start_date, '%Y-%m-%d')
        datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        print("날짜 형식이 잘못되었습니다.")
        return

    print()
    print(f"📅 대상 기간: {start_date} ~ {end_date}")
    print()

    # 0. 데이터 저장소 초기화
    data_store = ClassifiedDataStore(
        classified_data_dir=os.getenv("CLASSIFIED_DATA_DIR", "./data/classified_data"),
        stats_dir=os.getenv("STATS_DIR", "./data/stats")
    )

    # 1. 저장된 분류 결과 확인
    print("1️⃣  분류 데이터 확인")
    print("-" * 60)

    if data_store.exists(start_date):
        print(f"✓ 저장된 분류 결과 발견!")
        print(f"  로드 중...")

        classified_data = data_store.load_weekly_data(start_date)
        classified_letters = classified_data['letters']
        classified_posts = classified_data['posts']

        print(f"✓ 편지글 {len(classified_letters)}건 로드")
        print(f"✓ 게시글 {len(classified_posts)}건 로드")
        print(f"⚡ 재분류 생략 (기존 데이터 재사용)")
    else:
        print(f"❌ 저장된 분류 결과 없음")
        print(f"  BigQuery 조회 및 분류 시작...")
        print()

        # BigQuery 데이터 조회
        print("  📊 BigQuery 데이터 조회")
        print("  " + "-" * 58)

        client = BigQueryClient()
        query_with_client = WeeklyDataQuery(client)

        # 마스터 정보 조회
        print("  마스터 정보 조회 중...")
        master_info = query_with_client.get_master_info()
        print(f"  ✓ {len(master_info)}개 마스터/게시판 정보 로드")

        weekly_data = query_with_client.get_weekly_data(start_date, end_date)
        letters = weekly_data['letters']
        posts = weekly_data['posts']

        print(f"  ✓ 편지글 {len(letters)}건 조회")
        print(f"  ✓ 게시글 {len(posts)}건 조회")
        print()

        if not letters and not posts:
            print("  ❌ 데이터가 없어 리포트를 생성할 수 없습니다.")
            return

        # 게시판 -> 마스터 매핑 조회
        board_to_master_query = f"""
        SELECT _id as boardId, masterId
        FROM `{client.project_id}.us_plus.postboards`
        """
        board_to_master = {b['boardId']: b['masterId']
                           for b in client.execute_query(board_to_master_query)}

        # 편지글: 마스터 이름 추가
        for item in letters:
            master_id = item.get('masterId')
            if master_id and master_id in master_info:
                item['masterName'] = master_info[master_id]['displayName']
                item['masterClubName'] = master_info[master_id]['clubName']
                item['actualMasterId'] = master_id
            else:
                item['masterName'] = 'Unknown'
                item['masterClubName'] = 'Unknown'
                item['actualMasterId'] = master_id or 'unknown'

        # 게시글: postBoardId를 실제 masterId로 변환
        for item in posts:
            board_id = item.get('postBoardId')
            actual_master_id = board_to_master.get(board_id, board_id)

            if actual_master_id and actual_master_id in master_info:
                item['masterName'] = master_info[actual_master_id]['displayName']
                item['masterClubName'] = master_info[actual_master_id]['clubName']
                item['actualMasterId'] = actual_master_id
            else:
                item['masterName'] = 'Unknown'
                item['masterClubName'] = 'Unknown'
                item['actualMasterId'] = actual_master_id or 'unknown'

        # 서비스 공지글 필터링
        print("  🔍 서비스 공지글 필터링")
        letters = filter_service_notices(letters, content_field="message")
        posts = filter_service_notices(posts, content_field="textBody")
        print()

        # 콘텐츠 분류 (벡터 기반)
        print("  📝 콘텐츠 분류 (벡터 유사도 기반)")
        print("  " + "-" * 58)

        classifier = VectorContentClassifier()

        if letters:
            print(f"  편지글 {len(letters)}건 분류 중...")
            classified_letters = classifier.classify_batch(
                letters,
                content_field="message"
            )
            print(f"  ✓ 편지글 분류 완료")
        else:
            classified_letters = []

        if posts:
            print(f"  게시글 {len(posts)}건 분류 중...")
            classified_posts = classifier.classify_batch(
                posts,
                content_field="textBody"
            )
            print(f"  ✓ 게시글 분류 완료")
        else:
            classified_posts = []

        print()

        # 서비스 카테고리 LLM 후처리 검토
        print("  🔍 서비스 카테고리 LLM 후처리 검토")
        print("  " + "-" * 58)

        reviewer = ServiceCategoryReviewer()

        # 편지글 검토
        service_letters = [l for l in classified_letters if l.get("classification", {}).get("category") in ["서비스 문의", "서비스 불편"]]
        if service_letters:
            print(f"  편지글 중 서비스 카테고리 {len(service_letters)}건 검토 중...")
            classified_letters, letter_changes = reviewer.review_batch(classified_letters, content_field="message")
            if letter_changes:
                print(f"  ✓ 편지글 {len(letter_changes)}건 카테고리 변경됨")
                for change in letter_changes:
                    print(f"    - [{change['from']}] → [{change['to']}]: {change['content'][:50]}...")

        # 게시글 검토
        service_posts = [p for p in classified_posts if p.get("classification", {}).get("category") in ["서비스 문의", "서비스 불편"]]
        if service_posts:
            print(f"  게시글 중 서비스 카테고리 {len(service_posts)}건 검토 중...")
            classified_posts, post_changes = reviewer.review_batch(classified_posts, content_field="textBody")
            if post_changes:
                print(f"  ✓ 게시글 {len(post_changes)}건 카테고리 변경됨")
                for change in post_changes:
                    print(f"    - [{change['from']}] → [{change['to']}]: {change['content'][:50]}...")

        print()

        # 분류 결과 저장 (2-Tier)
        print("  💾 분류 결과 저장 (2-Tier)")
        print("  " + "-" * 58)

        data_store.save_weekly_data(
            start_date,
            end_date,
            classified_letters,
            classified_posts
        )

        print(f"  ✓ 전체 데이터 저장: data/classified_data/{start_date}.json")
        print(f"  ✓ 통계 요약 저장: data/stats/{start_date}.json")

    print()

    # 2. 벡터 스토어에 저장
    print("2️⃣  벡터 스토어 저장")
    print("-" * 60)

    try:
        store = ChromaVectorStore(
            collection_name=f"week_{start_date}",
            persist_directory="./chroma_db"
        )

        # 기존 데이터 초기화
        store.reset()

        # 데이터 저장
        total_added = 0
        if classified_letters:
            for letter in classified_letters:
                letter["message"] = letter.get("message", "")
            added = store.add_contents_batch(classified_letters, text_field="message")
            total_added += added

        if classified_posts:
            for post in classified_posts:
                post["message"] = post.get("textBody") or post.get("body", "")
            added = store.add_contents_batch(classified_posts, text_field="message")
            total_added += added

        print(f"✓ {total_added}건 벡터 스토어에 저장 완료")
    except Exception as e:
        print(f"⚠️  벡터 스토어 저장 실패: {str(e)}")

    print()

    # 3. 전주 데이터 로드 (전주 비교)
    print("3️⃣  전주 데이터 로드")
    print("-" * 60)

    prev_start, prev_end = WeeklyDataQuery.get_previous_week_range()
    print(f"📅 전주 기간: {prev_start} ~ {prev_end}")

    previous_letters = None
    previous_posts = None

    if data_store.exists(prev_start):
        try:
            previous_data = data_store.load_weekly_data(prev_start)
            previous_letters = previous_data['letters']
            previous_posts = previous_data['posts']
            print(f"✓ 전주 데이터 로드: 편지 {len(previous_letters)}건, 게시글 {len(previous_posts)}건")
        except Exception as e:
            print(f"⚠️  전주 데이터 로드 실패: {str(e)}")
    else:
        print(f"❌ 전주 데이터 없음")

    print()

    # 4. 통계 분석
    print("4️⃣  통계 분석 (전주 비교)")
    print("-" * 60)

    analytics = WeeklyAnalytics()
    stats = analytics.analyze_weekly_data(
        classified_letters,
        classified_posts,
        previous_letters=previous_letters,
        previous_posts=previous_posts
    )

    total = stats["total_stats"]["this_week"]
    print(f"✓ 전체 통계: 편지 {total['letters']}건, 게시글 {total['posts']}건")

    category_stats = stats["category_stats"]
    print(f"✓ 카테고리별 통계:")
    for category, count in sorted(category_stats.items(), key=lambda x: x[1], reverse=True):
        print(f"  - {category}: {count}건")

    print()

    # 5. 엑셀 파일 생성
    print("5️⃣  엑셀 파일 생성")
    print("-" * 60)

    excel_dir = "./exports"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/labeling_test_{start_date}.xlsx"

    create_labeled_excel(classified_letters, classified_posts, excel_path)
    print()

    # 6. Slack 전송
    print("6️⃣  Slack 전송")
    print("-" * 60)

    slack = SlackNotifier()

    # 메인 메시지
    main_message = f"[라벨링 테스트] {start_date} ~ {end_date} 분류 결과 (편지 {len(classified_letters)}건, 게시글 {len(classified_posts)}건)"
    main_response = slack._send_message(main_message)

    if not main_response.get("ok"):
        print(f"메시지 전송 실패: {main_response.get('error')}")
        return

    message_ts = main_response.get("ts")
    print(f"✓ 메인 메시지 전송 완료")

    # 엑셀 파일 업로드
    print(f"  엑셀 파일 업로드 중...")
    upload_result = slack.upload_file_to_thread(
        file_path=excel_path,
        thread_ts=message_ts,
        title=f"라벨링 테스트 ({start_date})",
        comment=""
    )

    if upload_result.get("ok"):
        print(f"✓ 업로드 완료: {upload_result.get('file_url')}")
    else:
        print(f"❌ 업로드 실패: {upload_result.get('error')}")

    print()
    print("=" * 60)
    print("✅ 라벨링 테스트 완료!")
    print("=" * 60)


if __name__ == "__main__":
    main()
