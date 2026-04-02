"""최종 엑셀 출력 — 개발팀용 + 비개발팀용

개발팀용: 기술이슈 + 기능요청 클러스터 (세분류 포함)
비개발팀용: 운영이슈 + CS이슈 원본 데이터 (클러스터링 없이)
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "./exports"

# 스타일
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)


def styled_header(ws, headers, row=1):
    for col, name in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def styled_row(ws, row_idx, values, wrap_cols=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in (wrap_cols or [])))


def load_data(period):
    """클러스터링 엑셀 + 검수 JSON 로드"""
    from openpyxl import load_workbook

    excel_path = f"exports/tech_issues_clustered_2026_{period}.xlsx"
    review_path = f"exports/tech_issues_reviewed_2026_{period}.json"

    wb = load_workbook(excel_path)

    # 상세 데이터
    ws_detail = wb["전체 상세"]
    detail_rows = []
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        detail_rows.append({
            'cluster': row[0], 'desc': row[1], 'subtag': row[2],
            'source': row[3], 'master': row[4], 'sentiment': row[5],
            'text': row[6], 'date': row[7],
        })

    # Jira Tasks
    ws_tasks = wb["Jira Tasks"]
    task_rows = {}
    for row in ws_tasks.iter_rows(min_row=2, values_only=True):
        if row[0]:
            task_rows[row[0]] = {
                'count': row[1], 'rep_desc': row[2], 'communities': row[3],
                'sentiment_dist': row[4], 'date_range': row[5],
                'subtag_dist': row[6], 'quote1': row[7], 'quote2': row[8],
                'jira_title': row[9], 'jira_desc': row[10],
            }

    # 검수 결과
    with open(review_path) as f:
        review = json.load(f)

    return detail_rows, task_rows, review


def get_month_label(period):
    if period == '01':
        return '1월'
    elif period == '02_03':
        return '2~3월'
    return period


def export_dev_excel(periods):
    """개발팀용: 기술이슈 + 기능요청 클러스터"""
    wb = Workbook()
    wb.remove(wb.active)

    for period in periods:
        detail_rows, task_rows, review = load_data(period)
        clusters = review['clusters']
        month = get_month_label(period)

        # 기술이슈 + 기능요청 클러스터만
        tech_clusters = {
            cid: info for cid, info in clusters.items()
            if info.get('type', '') in ('기술이슈', '기능요청')
            or '기술이슈' in info.get('type', '')
            or '기능요청' in info.get('type', '')
        }

        # ── 이슈 요약 시트 ──
        ws1 = wb.create_sheet(f"{month} 이슈 요약")
        headers = ['유형', '클러스터 라벨', '건수', '영향 커뮤니티', '날짜 범위', '대표 원문']
        styled_header(ws1, headers)

        row_idx = 2
        for cid, info in sorted(tech_clusters.items(), key=lambda x: -x[1].get('item_count', 0)):
            task = task_rows.get(cid, {})
            values = [
                info.get('type', ''),
                info.get('label', ''),
                info.get('item_count', 0),
                task.get('communities', ''),
                task.get('date_range', ''),
                str(task.get('quote1', ''))[:250],
            ]
            styled_row(ws1, row_idx, values, wrap_cols=[6])
            row_idx += 1

        for i, w in enumerate([10, 45, 8, 25, 22, 50], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"
        ws1.freeze_panes = 'A2'

        # ── 상세 시트 ──
        ws2 = wb.create_sheet(f"{month} 상세")
        headers2 = ['유형', '클러스터 라벨', '출처', '마스터', '감성', '원문', '작성일']
        styled_header(ws2, headers2)

        row_idx = 2
        for cid, info in sorted(tech_clusters.items(), key=lambda x: -x[1].get('item_count', 0)):
            items = [d for d in detail_rows if d['cluster'] == cid]
            for item in items:
                values = [
                    info.get('type', ''),
                    info.get('label', ''),
                    item.get('source', ''),
                    item.get('master', ''),
                    item.get('sentiment', ''),
                    str(item.get('text', ''))[:300],
                    item.get('date', ''),
                ]
                styled_row(ws2, row_idx, values, wrap_cols=[6])
                row_idx += 1

        for i, w in enumerate([10, 40, 8, 10, 8, 50, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row_idx-1}"
        ws2.freeze_panes = 'A2'

    path = os.path.join(OUTPUT_DIR, "Q1_2026_기술이슈_개발팀.xlsx")
    wb.save(path)
    print(f"개발팀용: {path}")
    return path


def export_ops_excel(periods):
    """비개발팀용: 운영이슈 + CS이슈 (클러스터 라벨 포함, 세분류 없음)"""
    wb = Workbook()
    wb.remove(wb.active)

    for period in periods:
        detail_rows, task_rows, review = load_data(period)
        clusters = review['clusters']
        month = get_month_label(period)

        # 운영이슈 + CS이슈
        ops_clusters = {
            cid: info for cid, info in clusters.items()
            if info.get('type', '') in ('운영이슈', 'CS이슈')
            or '운영이슈' in info.get('type', '')
            or 'CS이슈' in info.get('type', '')
        }

        # ── 이슈 요약 시트 ──
        ws1 = wb.create_sheet(f"{month} 이슈 요약")
        headers = ['유형', '이슈 내용', '건수', '영향 커뮤니티', '날짜 범위']
        styled_header(ws1, headers)

        row_idx = 2
        for cid, info in sorted(ops_clusters.items(), key=lambda x: -x[1].get('item_count', 0)):
            task = task_rows.get(cid, {})
            values = [
                info.get('type', ''),
                info.get('label', ''),
                info.get('item_count', 0),
                task.get('communities', ''),
                task.get('date_range', ''),
            ]
            styled_row(ws1, row_idx, values)
            row_idx += 1

        for i, w in enumerate([10, 50, 8, 25, 22], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"
        ws1.freeze_panes = 'A2'

        # ── 원문 시트 ──
        ws2 = wb.create_sheet(f"{month} 원문")
        headers2 = ['유형', '이슈 내용', '출처', '마스터', '감성', '원문', '작성일']
        styled_header(ws2, headers2)

        row_idx = 2
        for cid, info in sorted(ops_clusters.items(), key=lambda x: -x[1].get('item_count', 0)):
            items = [d for d in detail_rows if d['cluster'] == cid]
            for item in items:
                values = [
                    info.get('type', ''),
                    info.get('label', ''),
                    item.get('source', ''),
                    item.get('master', ''),
                    item.get('sentiment', ''),
                    str(item.get('text', ''))[:300],
                    item.get('date', ''),
                ]
                styled_row(ws2, row_idx, values, wrap_cols=[6])
                row_idx += 1

        for i, w in enumerate([10, 40, 8, 10, 8, 50, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row_idx-1}"
        ws2.freeze_panes = 'A2'

    path = os.path.join(OUTPUT_DIR, "Q1_2026_운영CS이슈_비개발팀.xlsx")
    wb.save(path)
    print(f"비개발팀용: {path}")
    return path


if __name__ == "__main__":
    periods = ['01', '02_03']
    dev_path = export_dev_excel(periods)
    ops_path = export_ops_excel(periods)

    print(f"\n완료!")
    print(f"  개발팀: {dev_path}")
    print(f"  비개발팀: {ops_path}")
