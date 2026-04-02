"""클러스터 LLM 검수 — 노이즈 제거 + 라벨 보정

클러스터별로 원문을 LLM에 보여주고:
1. 이 클러스터의 적절한 한줄 라벨
2. 대응 필요 여부 (기술이슈/운영이슈/CS이슈/대응불필요)
3. 클러스터에 안 맞는 항목 인덱스
"""
import json
import time
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import boto3
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

import sys
_target = sys.argv[1] if len(sys.argv) > 1 else "02_03"
EXCEL_PATH = f"exports/tech_issues_clustered_2026_{_target}.xlsx"
OUTPUT_PATH = f"exports/tech_issues_reviewed_2026_{_target}.json"

REVIEW_PROMPT = """당신은 금융 교육 플랫폼의 VOC 클러스터 검수자입니다.

아래는 유사한 VOC를 묶은 하나의 클러스터입니다.

## 판단 기준

### 대응 유형
- 기술이슈: 앱/웹 버그, 접속 오류, 결제 시스템 오류, 알림 미작동 등 개발팀 대응 필요
- 기능요청: 새 기능 요청, UX 개선 요청 등 제품팀 검토 필요
- 운영이슈: 운영정책 관련, 콘텐츠 업로드, CS 응답 등 운영팀 확인 필요
- CS이슈: 환불/해지/구독 변경 등 고객서비스 대응 필요
- 대응불필요: 투자 담론, 감사/인사, 일상 잡담, 마스터에 대한 개인 의견, 외부 서비스(삼성증권/키움증권/유튜브 등) 이슈

### 주의사항
- **원문을 반드시 읽고 판단할 것**. 서술(description)이 아니라 원문 기준으로 유형 판단
- 투자 손실 불만, 마스터 비판, 시장 의견 등이 "접속 불가" 등으로 잘못 서술된 경우 → 대응불필요로 판정
- 삼성증권, 키움증권, 유튜브 등 외부 서비스 이슈는 대응불필요

### 검수 규칙
- 원문을 하나하나 읽고 판단하세요
- 클러스터에 안 맞는 항목이 있으면 인덱스를 명시하세요
- 라벨은 구체적으로 (예: "줌 라이브 접속 불가" O, "접속 문제" X)
- **라벨은 원문에서 관찰되는 구체적 사실/현상만 기술** (평가, 인과 추정, 귀책 금지)
- ⚠️ 금지: "미흡", "부재", "부족", "지연", "부적절", "관리", "대응" 등 평가/귀책 단어
- ⚠️ 금지: "~로 인한", "~때문에" 등 인과관계 추정
  - X: "운영진 커뮤니티 관리 부재"
  - O: "게시판 내 특정 글 지속 노출에 대한 사용자 제보"
  - X: "CS팀 응대 지연"
  - O: "문의 후 7일 이상 답변 미수신 보고"
  - X: "접근 권한 관리 미흡으로 혼란"
  - O: "미구독 사용자의 커뮤니티 접근이 가능한 상태라는 보고"
  - X: "자동결제 해제 안내 부재로 의도치 않은 결제"
  - O: "자동결제가 되는 줄 몰랐다는 사용자의 환불 요청"
- 인사이트/판단은 사람이 할 영역. 라벨은 **원문에서 보이는 사실만**
- ⚠️ "~로 보입니다", "~로 판단됩니다", "VOC가 아닙니다" 등 메타 코멘트를 라벨에 넣지 말 것

## 응답: JSON만 출력
{"label": "줌 라이브 접속 불가", "type": "기술이슈", "misplaced": [2, 5], "note": "2번은 투자 의견, 5번은 단순 인사"}"""


def review_clusters():
    wb = load_workbook(EXCEL_PATH)
    ws_tasks = wb["Jira Tasks"]
    ws_detail = wb["전체 상세"]

    # 클러스터별 데이터 구성
    detail_rows = list(ws_detail.iter_rows(min_row=2, values_only=True))
    # headers: 클러스터, 증상서술, subtag, 출처, 마스터, 감성, 원문, 작성일

    clusters = {}
    for row in detail_rows:
        cid = row[0]
        if cid not in clusters:
            clusters[cid] = []
        clusters[cid].append({
            'desc': row[1],
            'subtag': row[2],
            'source': row[3],
            'master': row[4],
            'sentiment': row[5],
            'text': row[6],
            'date': row[7],
        })

    # task 정보
    task_info = {}
    for row in ws_tasks.iter_rows(min_row=2, values_only=True):
        if row[0]:
            task_info[row[0]] = {'count': row[1], 'desc': row[2]}

    logger.info(f"클러스터 {len(clusters)}개, 총 {sum(len(v) for v in clusters.values())}건")

    # 2건 이상 클러스터만 검수 (1건짜리는 개별건이라 검수 의미 없음)
    review_targets = {k: v for k, v in clusters.items() if len(v) >= 2}
    logger.info(f"검수 대상: {len(review_targets)}개 클러스터 (2건 이상)")

    bedrock = boto3.client("bedrock-runtime", region_name="us-west-2")
    model_id = "us.anthropic.claude-haiku-4-5-20251001-v1:0"

    def review_single(cid, items):
        # 클러스터 원문 구성
        lines = [f"클러스터: {cid} ({len(items)}건)\n"]
        for i, item in enumerate(items):
            text = str(item.get('text', ''))[:200]
            lines.append(f"[{i}] ({item.get('source','')}, {item.get('master','')}, {item.get('date','')}) {text}")

        user_msg = '\n'.join(lines)

        for attempt in range(3):
            try:
                resp = bedrock.invoke_model(
                    modelId=model_id,
                    body=json.dumps({
                        "anthropic_version": "bedrock-2023-05-31",
                        "max_tokens": 300,
                        "system": REVIEW_PROMPT,
                        "messages": [{"role": "user", "content": user_msg[:3000]}],
                    }),
                )
                result = json.loads(resp["body"].read())
                raw = result["content"][0]["text"].strip()

                start = raw.find("{")
                end = raw.rfind("}") + 1
                if start >= 0 and end > start:
                    parsed = json.loads(raw[start:end])
                    # 빈 라벨 fallback
                    if not parsed.get('label', '').strip():
                        parsed['label'] = f"클러스터 {cid} ({len(items)}건)"
                    return parsed
                return {"label": f"파싱실패 ({cid})", "type": "대응불필요", "misplaced": [], "note": raw[:100]}
            except Exception as e:
                if "Throttl" in str(e) and attempt < 2:
                    time.sleep(3 * (attempt + 1))
                    continue
                return {"label": "검수실패", "type": "대응불필요", "misplaced": [], "note": str(e)[:100]}

    # 병렬 검수
    results = {}
    done = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=15) as executor:
        future_map = {}
        for cid, items in review_targets.items():
            future = executor.submit(review_single, cid, items)
            future_map[future] = cid

        for future in as_completed(future_map):
            cid = future_map[future]
            review = future.result()
            results[cid] = review
            done += 1
            if done % 20 == 0 or done == len(review_targets):
                logger.info(f"  검수 {done}/{len(review_targets)}")

    elapsed = time.time() - start_time
    logger.info(f"검수 완료: {elapsed:.1f}초")

    # 1건짜리는 검수 없이 기본값
    for cid, items in clusters.items():
        if cid not in results:
            text = str(items[0].get('text', ''))[:50]
            results[cid] = {
                "label": str(items[0].get('desc', text)),
                "type": "미검수",
                "misplaced": [],
                "note": "1건 클러스터 — 개별 모니터링"
            }

    # 통계
    type_counts = {}
    for r in results.values():
        t = r.get('type', '미분류')
        type_counts[t] = type_counts.get(t, 0) + 1

    print(f"\n{'='*60}")
    print(f"검수 결과")
    print(f"{'='*60}")
    for t, c in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}개 클러스터")

    misplaced_total = sum(len(r.get('misplaced', [])) for r in results.values())
    print(f"  오분류 항목: {misplaced_total}건")

    # 저장
    output = {
        'clusters': {k: {**results[k], 'item_count': len(clusters[k])} for k in results},
        'stats': type_counts,
        'misplaced_total': misplaced_total,
    }
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    logger.info(f"저장: {OUTPUT_PATH}")

    return output


if __name__ == "__main__":
    review_clusters()
