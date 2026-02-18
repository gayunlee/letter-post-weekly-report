"""리포트 링크를 Slack에 보내고, 라벨링된 엑셀 파일을 스레드에 첨부하는 스크립트"""
import sys
import os
import json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from src.integrations.slack_client import SlackNotifier
from src.integrations.notion_client import NotionReportClient
from src.bigquery.queries import WeeklyDataQuery


def create_labeled_excel(start_date: str, output_path: str) -> str:
    """
    분류된 데이터를 엑셀 파일로 생성

    Args:
        start_date: 시작 날짜 (YYYY-MM-DD)
        output_path: 출력 파일 경로

    Returns:
        생성된 파일 경로
    """
    # 분류된 데이터 로드
    data_path = f"./data/classified_data/{start_date}.json"
    if not os.path.exists(data_path):
        raise FileNotFoundError(f"분류된 데이터가 없습니다: {data_path}")

    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    letters = data.get('letters', [])
    posts = data.get('posts', [])

    print(f"데이터 로드: 편지 {len(letters)}건, 게시글 {len(posts)}건")

    # 엑셀 워크북 생성
    wb = Workbook()

    # 스타일 정의
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # === 편지 시트 ===
    ws_letters = wb.active
    ws_letters.title = "편지"

    # 헤더
    letter_headers = ["마스터", "오피셜클럽", "내용", "카테고리", "생성일"]
    for col, header in enumerate(letter_headers, 1):
        cell = ws_letters.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 데이터
    for row_idx, item in enumerate(letters, 2):
        classification = item.get('classification', {})
        for col in range(1, 6):
            cell = ws_letters.cell(row=row_idx, column=col)
            cell.alignment = wrap_alignment

        ws_letters.cell(row=row_idx, column=1, value=item.get('masterName', ''))
        ws_letters.cell(row=row_idx, column=2, value=item.get('masterClubName', ''))
        ws_letters.cell(row=row_idx, column=3, value=(item.get('message', '') or '')[:1000])
        ws_letters.cell(row=row_idx, column=4, value=classification.get('category', ''))
        ws_letters.cell(row=row_idx, column=5, value=item.get('createdAt', '')[:10] if item.get('createdAt') else '')

    # 열 너비 조정
    ws_letters.column_dimensions['A'].width = 12
    ws_letters.column_dimensions['B'].width = 15
    ws_letters.column_dimensions['C'].width = 80
    ws_letters.column_dimensions['D'].width = 15
    ws_letters.column_dimensions['E'].width = 12

    # === 게시글 시트 ===
    ws_posts = wb.create_sheet(title="게시글")

    # 헤더
    post_headers = ["마스터", "오피셜클럽", "제목", "내용", "카테고리", "생성일"]
    for col, header in enumerate(post_headers, 1):
        cell = ws_posts.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 데이터
    for row_idx, item in enumerate(posts, 2):
        classification = item.get('classification', {})
        content = item.get('textBody') or item.get('body', '') or ''
        for col in range(1, 7):
            cell = ws_posts.cell(row=row_idx, column=col)
            cell.alignment = wrap_alignment

        ws_posts.cell(row=row_idx, column=1, value=item.get('masterName', ''))
        ws_posts.cell(row=row_idx, column=2, value=item.get('masterClubName', ''))
        ws_posts.cell(row=row_idx, column=3, value=(item.get('title', '') or '')[:200])
        ws_posts.cell(row=row_idx, column=4, value=content[:1000])
        ws_posts.cell(row=row_idx, column=5, value=classification.get('category', ''))
        ws_posts.cell(row=row_idx, column=6, value=item.get('createdAt', '')[:10] if item.get('createdAt') else '')

    # 열 너비 조정
    ws_posts.column_dimensions['A'].width = 12
    ws_posts.column_dimensions['B'].width = 15
    ws_posts.column_dimensions['C'].width = 40
    ws_posts.column_dimensions['D'].width = 80
    ws_posts.column_dimensions['E'].width = 15
    ws_posts.column_dimensions['F'].width = 12

    # 저장
    wb.save(output_path)
    print(f"엑셀 파일 생성: {output_path}")

    return output_path


def generate_summary(start_date: str) -> str:
    """통계 데이터 기반 3줄 요약 생성"""
    stats_path = f"./data/stats/{start_date}.json"
    if not os.path.exists(stats_path):
        return (
            "이번 주 전체 이용자 반응 규모는 전주 대비 변화가 있었습니다.\n"
            "편지와 게시글 모두 다양한 반응이 나타났습니다.\n"
            "상세 내용은 리포트를 확인해주세요."
        )

    with open(stats_path, 'r', encoding='utf-8') as f:
        stats = json.load(f)

    total = stats.get("total_stats", {})
    this_week = total.get("this_week", {})
    change = total.get("change", {})
    categories = stats.get("category_stats", {})

    # 증감 트렌드
    if change.get("total", 0) > 0:
        trend = "증가"
    elif change.get("total", 0) < 0:
        trend = "감소"
    else:
        trend = "유지"

    # 가장 많은 카테고리
    top_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)[:2]

    lines = [
        f"이번 주 전체 이용자 반응 규모는 전주 대비 {trend}한 흐름으로 나타났습니다.",
        f"편지 {this_week.get('letters', 0)}건, 게시글 {this_week.get('posts', 0)}건이 접수되었습니다.",
    ]

    if len(top_cats) >= 2:
        lines.append(f"'{top_cats[0][0]}' 카테고리가 가장 많았으며, '{top_cats[1][0]}'가 그 뒤를 이었습니다.")
    elif len(top_cats) == 1:
        lines.append(f"'{top_cats[0][0]}' 카테고리가 가장 많았습니다.")
    return "\n".join(lines)


def main():
    """메인 실행"""
    # 날짜 설정
    start_date, end_date = WeeklyDataQuery.get_last_week_range()
    print(f"대상 기간: {start_date} ~ {end_date}")

    # 리포트 파일 경로
    report_path = f"./reports/weekly_report_{start_date}.md"
    if not os.path.exists(report_path):
        print(f"리포트 파일이 없습니다: {report_path}")
        return

    # 1. 노션에 리포트 업로드
    print("\n노션에 리포트 업로드 중...")
    notion = NotionReportClient()
    week_label = notion.get_week_number_korean(start_date)

    with open(report_path, 'r', encoding='utf-8') as f:
        report_content = f.read()

    title = f"[{week_label}] 오피셜클럽 별 편지·게시글 통합 분석"
    notion_result = notion.create_report_page(
        title=title,
        markdown_content=report_content,
        start_date=start_date,
        end_date=end_date
    )
    notion_url = notion_result.get("url")
    print(f"노션 업로드 완료: {notion_url}")

    # 2. 엑셀 파일 생성
    excel_dir = "./exports"
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = f"{excel_dir}/labeled_data_{start_date}.xlsx"

    create_labeled_excel(start_date, excel_path)

    # 3. Slack 알림 전송
    slack = SlackNotifier()

    print(f"\nSlack 메시지 전송 중...")

    # 메인 메시지 전송
    main_message = f"[{week_label}] 오피셜클럽 별 편지·게시글 통합 분석 공유"
    main_response = slack._send_message(main_message)

    if not main_response.get("ok"):
        print(f"메시지 전송 실패: {main_response.get('error')}")
        return

    message_ts = main_response.get("ts")
    print(f"메인 메시지 전송 완료 (ts: {message_ts})")

    # 노션 링크 댓글 (3줄 요약 포함)
    from datetime import datetime, timedelta
    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    actual_end = (end_dt - timedelta(days=1)).strftime('%m/%d')
    start_short = datetime.strptime(start_date, '%Y-%m-%d').strftime('%m/%d')

    # 통계 기반 3줄 요약 생성
    summary = generate_summary(start_date)
    thread_message = (
        f"<{notion_url}|오피셜클럽 별 편지·게시글 통합 분석 ({start_short}~{actual_end})>을 작성하여 공유드립니다.\n\n"
        f"{summary}"
    )
    slack._send_message(thread_message, thread_ts=message_ts)
    print("노션 링크 댓글 전송 완료")

    # 4. 엑셀 파일 업로드 (코멘트 없이)
    print(f"\n엑셀 파일 업로드 중...")
    upload_result = slack.upload_file_to_thread(
        file_path=excel_path,
        thread_ts=message_ts,
        title=f"분류 데이터 ({start_date})",
        comment=""
    )

    if upload_result.get("ok"):
        print(f"엑셀 파일 업로드 완료: {upload_result.get('file_url')}")
    else:
        print(f"업로드 실패: {upload_result.get('error')}")

    print("\n완료!")


if __name__ == "__main__":
    main()
