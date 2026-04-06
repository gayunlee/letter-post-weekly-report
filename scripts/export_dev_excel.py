"""개발팀용 기술이슈 엑셀 — 대카테고리/소카테고리/클러스터 라벨/subtag/생성일

월별 시트 분리, 1~3월
"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "exports/Q1_2026_개발팀_기술이슈.xlsx"

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=11)

# 대카테고리 매핑 (클러스터 라벨 키워드 기반)
def assign_big_category(label, issue_type):
    label_lower = label.lower()

    if issue_type == '기능요청':
        return '기능요청'

    # 접속/로그인
    if any(k in label for k in ['접속', '로그인', '입장', '진입']):
        return '접속/로그인'
    # 결제/구독
    if any(k in label for k in ['결제', '구독', '수강권', '등록', '미반영', '가입']):
        return '결제/구독'
    # 콘텐츠 표시
    if any(k in label for k in ['미노출', '미표시', '미업로드', '누락', '표시되지', '로딩']):
        return '콘텐츠 표시'
    # 알림
    if any(k in label for k in ['알림', '푸시', '알람']):
        return '알림'
    # 미디어/재생
    if any(k in label for k in ['재생', '음성', '영상', '음질', '화질', '오디오', '미디어']):
        return '미디어/재생'
    # 링크
    if any(k in label for k in ['링크', 'URL']):
        return '링크'
    # UI/UX
    if any(k in label for k in ['글자', '확대', '화면', '크기', 'UI', '네비게이션', '스크롤']):
        return 'UI/UX'

    return '기타'


def _suggest_action(big_cat, label, count, master_count):
    """대카테고리 + 라벨 기반 조치 권장 생성"""
    scope = '플랫폼 전체' if master_count >= 3 else '특정 커뮤니티'

    actions = {
        '접속/로그인': f'접속 안정성 점검, 에러 로그 모니터링 ({scope})',
        '결제/구독': f'결제 프로세스 디버깅, 결제-수강권 연동 점검',
        '콘텐츠 표시': f'콘텐츠 로딩/렌더링 파이프라인 점검',
        '알림': f'FCM/APNs 토큰 갱신 로직 점검, 알림 발송 성공/실패 로그 모니터링',
        '미디어/재생': f'미디어 플레이어 및 스트리밍 인프라 점검',
        '링크': f'링크 렌더링 및 리다이렉트 로직 점검',
        'UI/UX': f'UI 렌더링 및 반응형 레이아웃 점검',
        '기능요청': f'제품팀 검토 후 백로그 등록 권장',
    }

    base = actions.get(big_cat, '이슈 상세 확인 후 대응 판단')

    if count >= 8:
        base = f'[즉시 대응] {base}'
    elif count >= 4:
        base = f'[금주 내 확인] {base}'

    return base


def load_period(period_key):
    """클러스터 엑셀 + 검수 JSON 로드"""
    from openpyxl import load_workbook as lwb

    excel = Path(f"exports/tech_issues_clustered_2026_{period_key}.xlsx")
    review_path = Path(f"exports/tech_issues_reviewed_2026_{period_key}.json")

    if not excel.exists() or not review_path.exists():
        return [], {}

    wb = lwb(excel)
    ws = wb["전체 상세"]
    details = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        details.append({
            'cluster': row[0], 'desc': row[1], 'subtag': row[2],
            'source': row[3], 'master': row[4], 'sentiment': row[5],
            'text': row[6], 'date': str(row[7] or ''),
        })

    with open(review_path) as f:
        review = json.load(f)

    return details, review.get('clusters', {})


def main():
    wb = Workbook()
    wb.remove(wb.active)

    periods = [
        ('01', '1월'),
        ('02', '2월'),
        ('03', '3월'),
    ]

    for period_key, month_label in periods:
        details, clusters = load_period(period_key)
        if not details:
            continue

        # 기술이슈 + 기능요청 클러스터만
        tech_cids = {
            cid for cid, info in clusters.items()
            if info.get('type', '') in ('기술이슈', '기능요청')
            or '기술이슈' in info.get('type', '')
            or '기능요청' in info.get('type', '')
        }

        tech_details = [d for d in details if d['cluster'] in tech_cids]

        # ── 요약 시트 ──
        ws1 = wb.create_sheet(f"{month_label} 요약")
        headers1 = ['우선순위', '대카테고리', '유형', '이슈 라벨', '건수', '영향 커뮤니티', '기간', '조치 권장']
        for col, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 클러스터별 집계
        cluster_rows = []
        for cid in tech_cids:
            info = clusters.get(cid, {})
            items = [d for d in details if d['cluster'] == cid]
            if not items:
                continue

            label = info.get('label', '')
            issue_type = info.get('type', '')
            big_cat = assign_big_category(label, issue_type)
            masters = sorted(set(d['master'] for d in items if d.get('master')))
            dates = sorted(d['date'] for d in items if d.get('date'))
            neg_count = sum(1 for d in items if d.get('sentiment') == '부정')
            neg_ratio = neg_count / len(items) if items else 0

            # 우선순위 산정
            if len(items) >= 8 or (len(items) >= 5 and neg_ratio > 0.6):
                priority = '🔴 긴급'
            elif len(items) >= 4 or (len(items) >= 3 and neg_ratio > 0.5):
                priority = '🟡 중요'
            elif len(items) >= 2:
                priority = '🟢 모니터링'
            else:
                priority = '⚪ 참고'

            # 조치 권장 생성
            action = _suggest_action(big_cat, label, len(items), len(masters))

            cluster_rows.append({
                'priority': priority,
                'big_cat': big_cat,
                'type': issue_type,
                'label': label,
                'count': len(items),
                'masters': ', '.join(masters[:5]) + (f' 외 {len(masters)-5}개' if len(masters) > 5 else ''),
                'date_range': f"{dates[0]} ~ {dates[-1]}" if dates else '',
                'action': action,
            })

        # 우선순위 → 대카테고리 → 건수 순 정렬
        p_order = {'🔴 긴급': 0, '🟡 중요': 1, '🟢 모니터링': 2, '⚪ 참고': 3}
        cluster_rows.sort(key=lambda x: (p_order.get(x['priority'], 9), x['big_cat'], -x['count']))

        for i, row in enumerate(cluster_rows, 2):
            values = [row['priority'], row['big_cat'], row['type'], row['label'], row['count'], row['masters'], row['date_range'], row['action']]
            for col, val in enumerate(values, 1):
                cell = ws1.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=(col in [4, 8]))

        for i, w in enumerate([12, 14, 10, 45, 8, 25, 22, 40], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.auto_filter.ref = f"A1:H{len(cluster_rows)+1}"
        ws1.freeze_panes = 'A2'

        # ── 상세 시트 ──
        ws2 = wb.create_sheet(f"{month_label} 상세")
        headers2 = ['대카테고리', '유형', '이슈 라벨', 'subtag', '출처', '마스터', '감성', '원문', '생성일']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        row_idx = 2
        for cid in sorted(tech_cids, key=lambda c: (
            assign_big_category(clusters.get(c, {}).get('label', ''), clusters.get(c, {}).get('type', '')),
            -len([d for d in details if d['cluster'] == c])
        )):
            info = clusters.get(cid, {})
            items = [d for d in details if d['cluster'] == cid]
            if not items:
                continue

            label = info.get('label', '')
            issue_type = info.get('type', '')
            big_cat = assign_big_category(label, issue_type)

            for item in sorted(items, key=lambda x: x.get('date', '')):
                values = [
                    big_cat, issue_type, label, item.get('subtag', ''),
                    item.get('source', ''), item.get('master', ''),
                    item.get('sentiment', ''), str(item.get('text', ''))[:300],
                    item.get('date', ''),
                ]
                for col, val in enumerate(values, 1):
                    cell = ws2.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=(col == 8))
                row_idx += 1

        for i, w in enumerate([14, 10, 40, 12, 8, 10, 8, 50, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row_idx-1}"
        ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f"저장: {OUTPUT}")

    # 통계
    for period_key, month_label in periods:
        details, clusters = load_period(period_key)
        tech_cids = {cid for cid, info in clusters.items() if info.get('type','') in ('기술이슈','기능요청') or '기술이슈' in info.get('type','') or '기능요청' in info.get('type','')}
        tech_details = [d for d in details if d['cluster'] in tech_cids]
        print(f"  {month_label}: {len(tech_cids)}개 이슈, {len(tech_details)}건")


if __name__ == "__main__":
    main()
