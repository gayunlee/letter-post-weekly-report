"""VOC 대시보드 — 개발팀 / 비개발팀 뷰

로컬: streamlit run dashboard/voc_dashboard.py
Cloud Run: 컨테이너에서 `streamlit run dashboard/voc_dashboard.py --server.port=8080`

데이터 소스: BigQuery voc_labelled.letters_posts (일간 파이프라인 적재)
"""
import json
import os
from datetime import date, timedelta, datetime
from pathlib import Path

import streamlit as st
import pandas as pd
import plotly.express as px

st.set_page_config(page_title="VOC 대시보드", layout="wide", initial_sidebar_state="expanded")

EXPORTS_DIR = Path("exports")
PROJECT_ID = os.getenv("BIGQUERY_PROJECT_ID", "us-service-data")


# ══════════════════════════════════════
# 데이터 로드
# ══════════════════════════════════════

@st.cache_resource
def get_bq_client():
    from google.cloud import bigquery
    return bigquery.Client(project=PROJECT_ID)


@st.cache_data(ttl=600)   # 10분 캐시
def load_from_bq(start_date: str, end_date: str) -> pd.DataFrame:
    """voc_labelled.letters_posts 에서 분류된 데이터를 DataFrame 으로."""
    client = get_bq_client()
    query = f"""
    SELECT id, source_type, master_id, master_name, user_id,
           content, created_at, topic, subtag, sentiment, summary,
           confidence, pipeline_date
    FROM (
      SELECT *, ROW_NUMBER() OVER (PARTITION BY id ORDER BY classified_at DESC) as rn
      FROM `{PROJECT_ID}.voc_labelled.letters_posts`
      WHERE pipeline_date >= '{start_date}' AND pipeline_date <= '{end_date}'
    )
    WHERE rn = 1
    """
    df = client.query(query).to_dataframe()
    # 대시보드 호환 필드
    df["masterName"] = df["master_name"].fillna("Unknown")
    df["createdAt"] = df["created_at"]
    df["text"] = df["content"]
    return df


@st.cache_data
def load_clusters(excel_path, review_path):
    """로컬 클러스터링 결과 (exports/) — 개발팀 뷰 전용, 없으면 None 반환."""
    if not Path(excel_path).exists() or not Path(review_path).exists():
        return None, None, None

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)

    ws_tasks = wb["Jira Tasks"]
    tasks = []
    for row in ws_tasks.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tasks.append({
                'cluster': row[0], 'count': row[1], 'desc': row[2],
                'communities': row[3], 'sentiment': row[4],
                'date_range': row[5], 'subtag_dist': row[6],
            })
    tasks_df = pd.DataFrame(tasks)

    ws_detail = wb["전체 상세"]
    details = []
    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        details.append({
            'cluster': row[0], 'desc': row[1], 'subtag': row[2],
            'source': row[3], 'master': row[4], 'sentiment': row[5],
            'text': row[6], 'date': row[7],
        })
    detail_df = pd.DataFrame(details)

    with open(review_path) as f:
        review = json.load(f)
    review_map = review.get('clusters', {})

    return tasks_df, detail_df, review_map


# ── 사이드바 ──
st.sidebar.title("VOC 대시보드")

role = st.sidebar.radio("직군 선택", ["개발팀", "비개발팀"], index=1)

# 동적 기간 선택 — 프리셋 + 커스텀
PRESETS = {
    "최근 7일": 7,
    "최근 14일": 14,
    "최근 30일": 30,
    "커스텀": None,
}
preset = st.sidebar.selectbox("기간 프리셋", list(PRESETS.keys()), index=0)

today = date.today()
if PRESETS[preset] is not None:
    end_date = today
    start_date = today - timedelta(days=PRESETS[preset])
else:
    start_date = st.sidebar.date_input("시작일", today - timedelta(days=7))
    end_date = st.sidebar.date_input("종료일", today)

start_str = start_date.strftime("%Y-%m-%d")
end_str = end_date.strftime("%Y-%m-%d")
st.sidebar.caption(f"기간: {start_str} ~ {end_str}")

# BigQuery 조회
with st.spinner(f"BigQuery 조회 중 ({start_str} ~ {end_str})..."):
    try:
        raw_df = load_from_bq(start_str, end_str)
    except Exception as e:
        st.error(f"BigQuery 조회 실패: {e}")
        st.stop()

if raw_df.empty:
    st.warning(f"해당 기간({start_str} ~ {end_str})에 데이터가 없습니다. 일간 파이프라인 실행 여부 확인 필요.")
    st.stop()

st.sidebar.success(f"✓ {len(raw_df):,}건 로드")

# 클러스터링 데이터 (개발팀 뷰 전용 — 로컬 파일 있을 때만)
period_key = start_date.strftime("%m")
cluster_excel = EXPORTS_DIR / f"tech_issues_clustered_2026_{period_key}.xlsx"
review_json = EXPORTS_DIR / f"tech_issues_reviewed_2026_{period_key}.json"
tasks_df, detail_df, review_map = load_clusters(cluster_excel, review_json)
has_clusters = tasks_df is not None

# 검수 결과 매핑 (클러스터 데이터 있을 때만)
if has_clusters:
    cluster_type_map = {cid: info.get('type', '미검수') for cid, info in review_map.items()}
    cluster_label_map = {cid: info.get('label', '') for cid, info in review_map.items()}
    tasks_df['type'] = tasks_df['cluster'].map(cluster_type_map).fillna('미검수')
    tasks_df['label'] = tasks_df['cluster'].map(cluster_label_map).fillna(tasks_df['desc'])
    detail_df['type'] = detail_df['cluster'].map(cluster_type_map).fillna('미검수')
    detail_df['label'] = detail_df['cluster'].map(cluster_label_map).fillna(detail_df['desc'])


# ══════════════════════════════════════
# 개발팀 뷰
# ══════════════════════════════════════
if role == "개발팀":
    st.title(f"개발팀 — 기술이슈 & 기능요청 ({start_str} ~ {end_str})")

    if not has_clusters:
        st.info(
            "클러스터링 데이터(`exports/tech_issues_clustered_*.xlsx`)가 없습니다. "
            "피드백 클러스터링 결과가 아직 BigQuery 로 이관되지 않은 상태입니다. "
            "비개발팀 뷰에서 원문 검색으로 이슈 확인이 가능합니다."
        )
        st.stop()

    # 기술이슈 + 기능요청만
    tech_types = {'기술이슈', '기능요청'}
    tech_tasks = tasks_df[tasks_df['type'].apply(lambda x: any(t in str(x) for t in tech_types))]
    tech_detail = detail_df[detail_df['type'].apply(lambda x: any(t in str(x) for t in tech_types))]

    # 통계
    col1, col2, col3, col4 = st.columns(4)
    n_tech = len(tech_tasks[tech_tasks['type'].str.contains('기술이슈')])
    n_feat = len(tech_tasks[tech_tasks['type'].str.contains('기능요청')])
    col1.metric("기술이슈", f"{n_tech}개")
    col2.metric("기능요청", f"{n_feat}개")
    col3.metric("총 보고 건수", f"{tech_detail.shape[0]}건")
    col4.metric("영향 커뮤니티", f"{tech_detail['master'].nunique()}개")

    st.divider()

    # 필터
    st.sidebar.divider()
    type_filter = st.sidebar.multiselect(
        "유형", ['기술이슈', '기능요청'], default=['기술이슈', '기능요청']
    )
    min_count = st.sidebar.slider("최소 건수", 1, max(int(tech_tasks['count'].max()), 2), 2)

    filtered = tech_tasks[
        tech_tasks['type'].apply(lambda x: any(t in str(x) for t in type_filter)) &
        (tech_tasks['count'] >= min_count)
    ].sort_values('count', ascending=False)

    # 이슈 목록
    st.subheader(f"이슈 목록 ({len(filtered)}개)")

    for _, row in filtered.iterrows():
        cid = row['cluster']
        count = row['count']
        label = row['label']
        issue_type = row['type']
        communities = row.get('communities', '')
        date_range = row.get('date_range', '')

        badge = "🔴" if count >= 5 else "🟡" if count >= 3 else "🟢"
        type_badge = "🐛" if '기술이슈' in str(issue_type) else "💡"

        with st.expander(f"{badge} {type_badge} {label} ({count}건)", expanded=(count >= 8)):
            cols = st.columns([1, 1, 1])
            cols[0].write(f"**유형:** {issue_type}")
            cols[1].write(f"**커뮤니티:** {communities}")
            cols[2].write(f"**기간:** {date_range}")

            # 원문
            items = detail_df[detail_df['cluster'] == cid]
            for _, item in items.iterrows():
                border = '#dc3545' if item.get('sentiment') == '부정' else '#28a745' if item.get('sentiment') == '긍정' else '#6c757d'
                st.markdown(f"""
<div style="background:#f8f9fa;padding:10px;border-radius:5px;margin:4px 0;border-left:3px solid {border};">
<small><b>{item.get('source','')}</b> | {item.get('master','')} | {item.get('date','')}</small><br>
<span style="color:#333;">{str(item.get('text',''))[:300]}</span>
</div>""", unsafe_allow_html=True)

    # 전체 테이블
    st.divider()
    with st.expander("전체 데이터 테이블"):
        st.dataframe(
            filtered[['type', 'label', 'count', 'communities', 'date_range']],
            width='stretch', height=400,
        )


# ══════════════════════════════════════
# 비개발팀 뷰
# ══════════════════════════════════════
else:
    st.title(f"VOC 인사이트 대시보드 ({start_str} ~ {end_str})")

    # ── 전체 요약 ──
    total = len(raw_df)
    topic_counts = raw_df['topic'].value_counts() if 'topic' in raw_df.columns else pd.Series()
    sentiment_counts = raw_df['sentiment'].value_counts() if 'sentiment' in raw_df.columns else pd.Series()

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("전체 VOC", f"{total:,}건")
    col2.metric("피드백", f"{topic_counts.get('피드백', 0):,}건")
    neg_count = sentiment_counts.get('부정', 0)
    col3.metric("부정 의견", f"{neg_count:,}건", f"{neg_count/total*100:.1f}%")
    pos_count = sentiment_counts.get('긍정', 0)
    col4.metric("긍정 의견", f"{pos_count:,}건", f"{pos_count/total*100:.1f}%")

    st.divider()

    # ── 차트 ──
    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.subheader("topic 분포")
        if not topic_counts.empty:
            fig = px.pie(values=topic_counts.values, names=topic_counts.index, hole=0.4)
            fig.update_layout(height=300, margin=dict(t=20, b=20))
            st.plotly_chart(fig, use_container_width=True)

    with chart_col2:
        st.subheader("감성 분포")
        if not sentiment_counts.empty:
            colors = {'긍정': '#28a745', '부정': '#dc3545', '중립': '#6c757d'}
            fig = px.bar(x=sentiment_counts.index, y=sentiment_counts.values,
                        color=sentiment_counts.index, color_discrete_map=colors)
            fig.update_layout(height=300, margin=dict(t=20, b=20), showlegend=False)
            st.plotly_chart(fig, use_container_width=True)

    # ── 마스터별 요약 ──
    st.divider()
    st.subheader("마스터별 요약")

    if 'masterName' in raw_df.columns:
        master_stats = raw_df.groupby('masterName').agg(
            total=('topic', 'count'),
            neg=('sentiment', lambda x: (x == '부정').sum()),
            pos=('sentiment', lambda x: (x == '긍정').sum()),
        ).sort_values('total', ascending=False).head(15)
        master_stats['부정비율'] = (master_stats['neg'] / master_stats['total'] * 100).round(1)

        st.dataframe(
            master_stats[['total', 'neg', 'pos', '부정비율']].rename(
                columns={'total': '전체', 'neg': '부정', 'pos': '긍정'}
            ),
            width='stretch',
        )

    # ── 주요 이슈 (운영/CS) — 클러스터 데이터 있을 때만 ──
    if has_clusters:
        st.divider()
        st.subheader("주요 이슈")

        ops_types = {'운영이슈', 'CS이슈'}
        ops_tasks = tasks_df[tasks_df['type'].apply(lambda x: any(t in str(x) for t in ops_types))]
        ops_tasks = ops_tasks.sort_values('count', ascending=False)

        for _, row in ops_tasks.head(15).iterrows():
            cid = row['cluster']
            count = row['count']
            label = row['label']
            issue_type = row['type']

            badge = "🔴" if count >= 5 else "🟡" if count >= 3 else "🟢"

            with st.expander(f"{badge} [{issue_type}] {label} ({count}건)"):
                cols = st.columns([1, 1])
                cols[0].write(f"**커뮤니티:** {row.get('communities', '')}")
                cols[1].write(f"**기간:** {row.get('date_range', '')}")

                items = detail_df[detail_df['cluster'] == cid]
                for _, item in items.iterrows():
                    border = '#dc3545' if item.get('sentiment') == '부정' else '#6c757d'
                    st.markdown(f"""
<div style="background:#f8f9fa;padding:10px;border-radius:5px;margin:4px 0;border-left:3px solid {border};">
<small><b>{item.get('source','')}</b> | {item.get('master','')} | {item.get('date','')}</small><br>
<span style="color:#333;">{str(item.get('text',''))[:300]}</span>
</div>""", unsafe_allow_html=True)

    # ── 원문 검색 ──
    st.divider()
    st.subheader("원문 검색")

    search_col1, search_col2, search_col3 = st.columns([2, 1, 1])
    keyword = search_col1.text_input("키워드", "")

    text_col = 'text' if 'text' in raw_df.columns else 'message'
    search_df = raw_df.copy()

    if 'topic' in search_df.columns:
        topic_filter = search_col2.multiselect("topic", raw_df['topic'].unique().tolist())
        if topic_filter:
            search_df = search_df[search_df['topic'].isin(topic_filter)]

    if 'masterName' in search_df.columns:
        master_filter = search_col3.multiselect("마스터", sorted(raw_df['masterName'].unique().tolist()))
        if master_filter:
            search_df = search_df[search_df['masterName'].isin(master_filter)]

    if keyword and text_col in search_df.columns:
        search_df = search_df[search_df[text_col].astype(str).str.contains(keyword, case=False, na=False)]

    st.write(f"**{len(search_df):,}건** 검색됨")

    # 결과 카드
    display = search_df.sort_values('createdAt', ascending=False).head(30) if 'createdAt' in search_df.columns else search_df.head(30)

    for _, item in display.iterrows():
        sentiment = item.get('sentiment', '중립')
        border = '#dc3545' if sentiment == '부정' else '#28a745' if sentiment == '긍정' else '#6c757d'
        topic = item.get('topic', '')
        subtag = item.get('subtag', '')
        master = item.get('masterName', '')
        date = str(item.get('createdAt', ''))[:10]
        text = str(item.get(text_col, ''))[:300]
        summary = item.get('summary', '')

        st.markdown(f"""
<div style="background:#f8f9fa;padding:12px;border-radius:5px;margin:6px 0;border-left:4px solid {border};">
<small><b>{topic}</b> > {subtag} | <b>{master}</b> | {date} | {sentiment}</small>
{f'<br><small style="color:#888;">요약: {summary}</small>' if summary else ''}
<div style="margin-top:6px;">{text}</div>
</div>""", unsafe_allow_html=True)
